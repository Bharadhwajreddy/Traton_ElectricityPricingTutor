#!/usr/bin/env python3
"""
Generate the Electricity Pricing Tutor Excel workbook.

Sheets: Master_Index, Inputs, Outputs, and one Country_<Name> sheet per
documented country (Germany, Sweden, China). The workbook mirrors the web app's
data model so formulas can later be filled with numeric tariff values.

Usage:  python3 scripts/build_excel.py
Output: public/electricity-pricing-tutor.xlsx
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "..", "public", "electricity-pricing-tutor.xlsx")

HEAD = Font(bold=True, color="FFFFFF")
HEAD_FILL = PatternFill("solid", fgColor="1D4ED8")
SUB = Font(bold=True, color="1E3A8A")
SUB_FILL = PatternFill("solid", fgColor="DBEAFE")
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Border(*[Side(style="thin", color="E2E8F0")] * 4)

# --- Country status registry (matches src/data) -----------------------------
MASTER = [
    ("Germany", "DE", "Europe", "deep-dive", "2026-06-11"),
    ("Sweden", "SE", "Europe", "deep-dive", "2026-06-11"),
    ("Denmark", "DK", "Europe", "deep-dive", "2026-06-15"),
    ("China", "CN", "Asia", "deep-dive", "2026-06-11"),
]
STUBS = ["France", "Netherlands", "Italy", "Spain", "United Kingdom", "Poland",
         "Belgium", "Austria", "Switzerland", "Norway", "Finland",
         "Portugal", "Ireland", "Czechia", "Slovakia", "Hungary", "Romania",
         "Bulgaria", "Greece", "Croatia", "Slovenia", "Estonia", "Latvia",
         "Lithuania", "Luxembourg", "Iceland", "Serbia", "Ukraine", "Malta",
         "Cyprus", "Albania", "Bosnia and Herzegovina", "North Macedonia",
         "Montenegro", "Kosovo", "Moldova"]

# --- Per-country content (compact mirror of the TS data) --------------------
# rows of (component, category, legal basis, calculation, region variation)
COUNTRIES = {
    "Germany": {
        "classification": "Medium voltage (10-20 kV), RLM-metered industrial Sondervertragskunde. "
                          "Utilisation t_B = E/P_max = 2000 h -> <2500 h Netzentgelt regime.",
        "components": [
            ("Energiepreis / Beschaffung", "energy", "Free contract (EPEX/EEX)", "E x p_energy (market)", "Supplier-specific"),
            ("Netzentgelt Leistungspreis", "grid-power", "§17 StromNEV; DSO Preisblatt", "P_max x LP(<2500h)", "Each DSO"),
            ("Netzentgelt Arbeitspreis", "grid-energy", "§17 StromNEV; DSO Preisblatt", "E x AP(<2500h)", "Each DSO"),
            ("Stromsteuer", "tax", "StromStG §9b", "E x 0.0205 EUR/kWh (=41,000)", "Federal"),
            ("Konzessionsabgabe", "concession", "§48 EnWG; KAV §2", "E x 0.0011 (cap; Grenzpreis test)", "Municipality"),
            ("KWKG-Umlage", "surcharge", "KWKG; EnFG", "E1 x 0.00446 + E2 x B'", "National"),
            ("§19 StromNEV-Umlage", "surcharge", "§19 StromNEV; EnFG", "E1 x 0.01559 + E2 x 0.00050 (=16,090)", "National"),
            ("Offshore-Netzumlage", "surcharge", "§17f EnWG; EnFG", "E1 x 0.00941 + E2 x B'", "National"),
            ("AbLaV-Umlage", "surcharge", "§18 AbLaV", "0 (not levied since 2023)", "National"),
            ("EEG-Umlage", "surcharge", "EEG/EnFG", "0 (abolished 2022)", "N/A"),
            ("Messstellenbetrieb", "metering", "MsbG §32", "MsbG-capped EUR/yr (not modelled)", "Msb-specific"),
            ("Mehrwertsteuer (VAT)", "tax", "UStG", "19% on net (recoverable)", "Federal"),
        ],
        "formulas": [
            ("Energy", "C_energy = E * p_energy"),
            ("Grid", "C_grid = P_max*LP + E*AP   (regime <2500h)"),
            ("Electricity tax", "C_stromst = E*0.0205 = 41000"),
            ("Concession", "C_konz = E*0.0011 = 2200"),
            ("KWKG", "C_KWKG = E1*0.00446 + E2*Bprime"),
            ("§19 StromNEV", "C_19 = E1*0.01559 + E2*0.00050 = 16090"),
            ("Offshore", "C_offshore = E1*0.00941 + E2*Bprime"),
            ("TOTAL (net)", "C_net = sum(above) + C_metering ;  C_gross = C_net*1.19"),
        ],
        "v2g": "§118 Abs.6 EnWG: storage/bidirectional (V2G) charging exempt from double grid fees; "
               "via §21 EnFG treated as storage (0.000 ct/kWh surcharge). BNetzA MiSpeL from 2026-04, no 2nd meter.",
    },
    "Sweden": {
        "classification": "Hogspanning power-subscription (Vattenfall N4) for ~1 MW. Energy is competitive "
                          "Nord Pool spot in the depot's SE1-SE4 zone. Grid = fixed + effektavgift + seasonal overforing.",
        "components": [
            ("Spotpris / elhandelspris", "energy", "Nord Pool day-ahead (SE1-SE4)", "E x p_spot (zone-dependent)", "Supplier & zone"),
            ("Pa slag (markup)", "energy", "Contract", "small ore/kWh adder", "Supplier"),
            ("Fast avgift", "fixed", "Ellagen; Ei", "12 x 560 = 6,720 SEK", "DSO"),
            ("Effektavgift", "grid-power", "Ellagen; Ei", "12 x 54 x P_max = 648,000 SEK", "DSO"),
            ("Hoglasteffektavgift", "grid-power", "Ellagen; Ei", "12 x 0 x P_max (Vattenfall N4)", "DSO"),
            ("Overforingsavgift", "grid-energy", "Ellagen; Ei", "E_hog x 0.612 + E_lag x 0.244 SEK", "DSO/zone"),
            ("Energiskatt", "tax", "Lag 1994:1776", "E x 0.360 = 720,000 SEK", "National (north reduction)"),
            ("Industri 0.6 ore (N/A)", "tax", "Lag 1994:1776", "Depot does NOT qualify (logistics)", "National"),
            ("Moms (VAT)", "tax", "Mervardesskattelagen", "25% (recoverable)", "National"),
            ("Elcertifikat", "surcharge", "Lag 2003:113", "small, phasing out (~0)", "Supplier"),
        ],
        "formulas": [
            ("Energy", "C_energy = E * p_spot  (zone-dependent)"),
            ("Grid fixed", "C_fast = 12*560 = 6720"),
            ("Grid power", "C_effekt = 12*54*P_max = 648000"),
            ("Grid energy", "C_overf = E_hog*0.612 + E_lag*0.244"),
            ("Energy tax", "C_skatt = E*0.360 = 720000"),
            ("TOTAL", "Subtotal_excl_VAT = sum(above);  Total = 1.25*Subtotal (VAT recoverable)"),
        ],
        "v2g": "Export => producer/inmatning customer (natnytta ~5.9 ore/kWh + energy sold at ~spot). "
               "60-ore micro-production tax credit does NOT apply (needs <=100 A AND abolished 2026-01-01). "
               "No dedicated V2G tariff regime; not modelled.",
    },
    "China": {
        "classification": "10 kV, 工商业 C&I, two-part (>=315 kVA mandatory). Energy market-determined "
                          "(agent-purchase 0.3715 yuan/kWh Apr-2026, Jiangsu). Demand: capacity (kVA) or max-demand (kW) basis.",
        "components": [
            ("Agent-purchase energy", "energy", "NDRC 1439 (2021)", "0.3715 yuan/kWh (monthly, Jiangsu)", "Province/month"),
            ("Line-loss charge", "grid-energy", "NDRC 526", "0.0123 yuan/kWh", "Province"),
            ("T&D energy price", "grid-energy", "NDRC 526 (2023)", "0.1357 yuan/kWh @10kV", "Province"),
            ("Capacity charge", "grid-power", "NDRC 526; two-part", "S_kVA x 32 x 12 = 384,000 yuan", "Province"),
            ("Max-demand charge", "grid-power", "NDRC 526; two-part", "P_max x 51.2 x 12 (x0.9 if >=260 kWh/kVA)", "Province"),
            ("System operation cost", "surcharge", "Market rules", "0.0688 yuan/kWh", "Province"),
            ("Govt funds & surcharges", "levy", "National funds", "0.0294 yuan/kWh (incl. renewable 0.019)", "Province"),
            ("ToU peak/flat/valley", "energy", "Su DRC 426/2025", "peak +80%/valley -65%; 0.9149/0.6177/0.3762", "Province"),
            ("Critical-peak/deep-valley", "energy", "Su DRC; >=315 kVA", "+20% / -20%", "Few provinces"),
            ("VAT (13%)", "tax", "National", "included in published yuan/kWh", "National"),
            ("Power-factor adjustment", "surcharge", "Lilv method", "+/- vs PF 0.9 (per account)", "Standard varies"),
        ],
        "formulas": [
            ("Energy (ToU)", "E_peak*0.9149 + E_flat*0.6177 + E_valley*0.3762"),
            ("Demand (capacity)", "S_kVA*32*12 = 384000"),
            ("Demand (max-demand)", "P_max*51.2*12  (x0.9 if kWh/kVA>=260)"),
            ("Power factor", "PF_cost = +/- beta*(Energy+Demand)"),
            ("TOTAL", "Total = Energy + Demand + PF   (VAT included). Illustrative ~1.65M yuan/yr"),
        ],
        "v2g": "V2G (车网互动) pilot-stage: 718/2024 + 241/2025 (9 cities, 30 projects). No national feed-in "
               "tariff; compensation via local pilots/aggregators/market. Not modelled.",
    },
    "Denmark": {
        "classification": "MV/HV (10-60 kV) business (C/B customer). Energy = Nord Pool spot, zone DK1 "
                          "(west) or DK2 (east). TSO Energinet; DSO time-of-day tariff (Radius Tarifmodel 3.0).",
        "components": [
            ("Spotpris (energy)", "energy", "Nord Pool day-ahead DK1/DK2", "E x p_spot (zone-dependent)", "Supplier & zone"),
            ("Energinet systemtarif", "grid-energy", "Energinet tariff reg.", "0.072 DKK/kWh (2026)", "National"),
            ("Energinet nettarif", "grid-energy", "Energinet tariff reg.", "0.043 DKK/kWh (2026); +187 kr sub", "National"),
            ("DSO nettarif (Tarifmodel 3.0)", "grid-energy", "DSO; Forsyningstilsynet", "ToU lav/hoj/spids; winter higher", "Each DSO"),
            ("Elafgift (electricity tax)", "tax", "Lov om afgift af elektricitet", "~0.008 DKK/kWh (2026 EU min); biz refund", "National"),
            ("PSO-tarif", "levy", "Former Elforsyningsloven", "0 (abolished end-2022)", "N/A"),
            ("Moms (VAT)", "tax", "Momsloven", "25% (recoverable)", "National"),
        ],
        "formulas": [
            ("Energy", "C_energy = E * p_spot (DK1/DK2)"),
            ("Energinet", "C_energinet = E*0.115 + 187 = 230000 + sub"),
            ("DSO grid", "C_dso = sum_period(E_period * rate_period)"),
            ("Electricity tax", "C_elafgift = E*0.008 ~ 16000 (biz refund)"),
            ("TOTAL", "C_preVAT = energy+energinet+dso+elafgift; x1.25 VAT (recoverable)"),
        ],
        "v2g": "Export => producer/feed-in under Energinet+DSO; sold at ~spot; Energinet flexibility/balancing "
               "markets. Early V2G testbed (Parker) but no dedicated V2G tariff as of mid-2026; not modelled.",
    },
}


def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEAD
        cell.fill = HEAD_FILL
        cell.alignment = WRAP


def section_row(ws, row, text, span):
    ws.cell(row=row, column=1, value=text)
    for c in range(1, span + 1):
        ws.cell(row=row, column=c).fill = SUB_FILL
        ws.cell(row=row, column=c).font = SUB


def build():
    wb = Workbook()

    # --- Master_Index ---
    ws = wb.active
    ws.title = "Master_Index"
    ws.append(["Country", "Code", "Region", "Status", "Last updated", "Sheet"])
    style_header(ws, 1, 6)
    for name, code, region, status, upd in MASTER:
        ws.append([name, code, region, status, upd, f"Country_{name}"])
    for name in sorted(STUBS):
        ws.append([name, "", "Europe", "needs-research", "2026-06-11", ""])
    for col, w in zip("ABCDEF", (24, 8, 10, 16, 14, 22)):
        ws.column_dimensions[col].width = w

    # --- Inputs (with named ranges) ---
    wi = wb.create_sheet("Inputs")
    wi.append(["Parameter", "Value", "Unit", "Name"])
    style_header(wi, 1, 4)
    wi.append(["Annual energy", 2000000, "kWh", "Depot_E"])
    wi.append(["Peak / contracted power", 1000, "kW", "Depot_Pmax"])
    wi.append(["Billing months", 12, "months", "Depot_Months"])
    wi.append(["Utilisation hours", "=Depot_E/Depot_Pmax", "h/yr", "Depot_Hours"])
    for col, w in zip("ABCD", (26, 16, 10, 16)):
        wi.column_dimensions[col].width = w
    wb.defined_names.add(DefinedName("Depot_E", attr_text="Inputs!$B$2"))
    wb.defined_names.add(DefinedName("Depot_Pmax", attr_text="Inputs!$B$3"))
    wb.defined_names.add(DefinedName("Depot_Months", attr_text="Inputs!$B$4"))

    # --- Outputs ---
    wo = wb.create_sheet("Outputs")
    wo.append(["Country", "Regulated/illustrative annual cost", "Currency", "Notes"])
    style_header(wo, 1, 4)
    wo.append(["Germany", "Stromsteuer+§19+Offshore+KWKG+Konz (regulated part)", "EUR",
               "Energy & grid (DSO) market-specific; see sheet"])
    wo.append(["Sweden", "=Depot_Months*54*Depot_Pmax + Depot_E*0.360", "SEK",
               "Effektavgift + energiskatt; energy zone-dependent"])
    wo.append(["China", "=Depot_Pmax*32*Depot_Months", "CNY",
               "Capacity charge only; +ToU energy ~1.27M, total ~1.65M (Jiangsu)"])
    for col, w in zip("ABCD", (16, 44, 10, 46)):
        wo.column_dimensions[col].width = w

    # --- Country sheets ---
    for name, code, region, status, upd in MASTER:
        d = COUNTRIES[name]
        cs = wb.create_sheet(f"Country_{name}")
        cs.cell(row=1, column=1, value=f"{name} — truck-depot electricity tariff").font = Font(bold=True, size=14)
        r = 3
        section_row(cs, r, "A · Classification & assumptions", 5); r += 1
        cs.cell(row=r, column=1, value=d["classification"]).alignment = WRAP
        cs.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        cs.row_dimensions[r].height = 45; r += 2

        section_row(cs, r, "B · Tariff components", 5); r += 1
        cs.append([]) if False else None
        for i, h in enumerate(["Component", "Category", "Legal basis", "Calculation", "Region variation"], 1):
            cell = cs.cell(row=r, column=i, value=h); cell.font = HEAD; cell.fill = HEAD_FILL; cell.alignment = WRAP
        r += 1
        for comp in d["components"]:
            for i, v in enumerate(comp, 1):
                cs.cell(row=r, column=i, value=v).alignment = WRAP
                cs.cell(row=r, column=i).border = THIN
            r += 1
        r += 1

        section_row(cs, r, "C · Symbolic formulas", 5); r += 1
        for label, f in d["formulas"]:
            cs.cell(row=r, column=1, value=label).font = Font(bold=True)
            cs.cell(row=r, column=2, value=f).alignment = WRAP
            cs.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
            r += 1
        r += 1

        section_row(cs, r, "D · V2G / feed-in / export", 5); r += 1
        cs.cell(row=r, column=1, value=d["v2g"]).alignment = WRAP
        cs.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        cs.row_dimensions[r].height = 60; r += 2

        section_row(cs, r, "F · References — see the web app country page for full citation list", 5)
        for col, w in zip("ABCDE", (30, 14, 26, 40, 22)):
            cs.column_dimensions[col].width = w

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {os.path.relpath(OUT)}")


if __name__ == "__main__":
    build()
